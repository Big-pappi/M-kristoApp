"""
generate_tracker.py
--------------------
Generates M-Kristo-Tracker.xlsx: a progress tracker + payment summary
workbook, built from the editable data in `tracker_data.py`.

Run:
    python generate_tracker.py

Requires: openpyxl (see guide.md for setup, or:
    python -m venv .venv && ./.venv/bin/pip install openpyxl
    ./.venv/bin/python generate_tracker.py
)
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import tracker_data as D

OUTPUT = "M-Kristo-Tracker.xlsx"

# --------------------------------------------------------------------------- #
#  Brand palette (matches docs/mkristo_doc_style.py)
# --------------------------------------------------------------------------- #
TEAL = "0E7C8B"
TEAL_DARK = "0A5C68"
PURPLE = "4A2E6B"
LIGHT = "EEF4F5"
INK = "1F2933"
MUTED = "5B6B78"
WHITE = "FFFFFF"

GREEN = "1E8E3E"   # Done / Paid
AMBER = "B8860B"   # In Progress / Pending
RED = "C0392B"     # Blocked
GRAY = "9AA5AC"    # Not Started

HEADER_FILL = PatternFill("solid", fgColor=TEAL)
TITLE_FILL = PatternFill("solid", fgColor=TEAL_DARK)
STRIPE_FILL = PatternFill("solid", fgColor=LIGHT)
TOTAL_FILL = PatternFill("solid", fgColor=PURPLE)

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=WHITE, size=16)
BODY_FONT = Font(name="Calibri", color=INK, size=10.5)
MUTED_FONT = Font(name="Calibri", color=MUTED, size=9.5, italic=True)
TOTAL_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)

THIN = Side(style="thin", color="CBD6D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _autosize(ws: Worksheet, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _banner(ws: Worksheet, span, title, subtitle):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    ws.merge_cells(f"A2:{get_column_letter(span)}2")
    c1 = ws["A1"]
    c1.value = title
    c1.font = TITLE_FONT
    c1.fill = TITLE_FILL
    c1.alignment = LEFT
    c2 = ws["A2"]
    c2.value = subtitle
    c2.font = Font(name="Calibri", color=WHITE, size=10.5)
    c2.fill = TITLE_FILL
    c2.alignment = LEFT
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    for row in (1, 2):
        for col in range(1, span + 1):
            ws.cell(row=row, column=col).fill = TITLE_FILL


def _header_row(ws: Worksheet, row, headers):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 22


def _status_fill_rules(ws: Worksheet, col_letter, first_row, last_row):
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    rules = [
        ("Done", GREEN), ("Paid", GREEN),
        ("In Progress", AMBER), ("Pending", AMBER),
        ("Blocked", RED),
        ("Not Started", GRAY),
    ]
    for text, color in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{text}"'],
                       fill=PatternFill("solid", fgColor=color),
                       font=Font(color=WHITE, bold=True)),
        )


# --------------------------------------------------------------------------- #
#  Sheet 1: Overview
# --------------------------------------------------------------------------- #
def build_overview(wb: Workbook):
    ws = wb.active
    ws.title = "Overview"
    _autosize(ws, [26, 26])
    _banner(ws, 2, D.PROJECT_NAME + " — Tracker", f"Client: {D.CLIENT}  |  Developer: {D.DEVELOPER}, {D.COMPANY}")

    total_tasks = len(D.TASKS)
    done_tasks = sum(1 for t in D.TASKS if t["status"] == "Done")
    in_progress = sum(1 for t in D.TASKS if t["status"] == "In Progress")
    overall_pct = round(sum(t["percent_complete"] for t in D.TASKS) / total_tasks) if total_tasks else 0

    paid = sum(p["amount"] for p in D.PAYMENTS if p["status"] == "Paid")
    pending = D.TOTAL_PROJECT_COST - paid

    rows = [
        ("Generated on", date.today().isoformat()),
        ("", ""),
        ("Total tasks", total_tasks),
        ("Done", done_tasks),
        ("In progress", in_progress),
        ("Overall progress", f"{overall_pct}%"),
        ("", ""),
        ("Total project cost", f"{D.TOTAL_PROJECT_COST:,.0f} {D.CURRENCY}"),
        ("Paid so far", f"{paid:,.0f} {D.CURRENCY}"),
        ("Remaining", f"{pending:,.0f} {D.CURRENCY}"),
    ]
    r = 4
    for label, value in rows:
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = Font(bold=True, color=INK, size=11)
        vc.font = BODY_FONT
        lc.alignment = LEFT
        vc.alignment = LEFT
        if label:
            lc.border = BORDER
            vc.border = BORDER
            row_fill = PatternFill("solid", fgColor=LIGHT) if r % 2 == 0 else PatternFill(fill_type=None)
            lc.fill = row_fill
            vc.fill = PatternFill("solid", fgColor=LIGHT) if r % 2 == 0 else PatternFill(fill_type=None)
        r += 1

    ws.row_dimensions[3].height = 6
    ws.freeze_panes = "A4"

    # Pie chart: task status breakdown
    status_counts = {}
    for t in D.TASKS:
        status_counts[t["status"]] = status_counts.get(t["status"], 0) + 1

    chart_start = r + 2
    ws.cell(row=chart_start - 1, column=1, value="Task status breakdown").font = Font(bold=True, color=INK, size=11)
    for i, (status, count) in enumerate(status_counts.items()):
        ws.cell(row=chart_start + i, column=1, value=status).font = BODY_FONT
        ws.cell(row=chart_start + i, column=2, value=count).font = BODY_FONT

    chart = PieChart()
    chart.title = "Tasks by status"
    data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + len(status_counts) - 1)
    cats = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_start + len(status_counts) - 1)
    chart.add_data(data)
    chart.set_categories(cats)
    chart.height, chart.width = 8, 12
    ws.add_chart(chart, f"D{chart_start - 1}")


# --------------------------------------------------------------------------- #
#  Sheet 2: Progress Tracker
# --------------------------------------------------------------------------- #
def build_progress(wb: Workbook):
    ws = wb.create_sheet("Progress Tracker")
    headers = ["Phase", "Task", "Owner", "Status", "% Complete", "Start Date", "Due Date", "Notes"]
    _autosize(ws, [22, 34, 16, 14, 12, 14, 14, 34])
    _banner(ws, len(headers), D.PROJECT_NAME + " — Progress Tracker",
            "Update this via tracker_data.py and rerun generate_tracker.py")

    header_row = 4
    _header_row(ws, header_row, headers)

    r = header_row + 1
    for t in D.TASKS:
        vals = [
            t["phase"], t["task"], t["owner"], t["status"],
            t["percent_complete"] / 100,
            t["start_date"].isoformat() if t["start_date"] else "",
            t["due_date"].isoformat() if t["due_date"] else "",
            t["notes"],
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.alignment = LEFT if col in (1, 2, 3, 8) else CENTER
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = STRIPE_FILL
            if col == 5:
                cell.number_format = "0%"
        r += 1

    last_row = r - 1
    _status_fill_rules(ws, "D", header_row + 1, last_row)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:H{last_row}"


# --------------------------------------------------------------------------- #
#  Sheet 3: Payment Summary
# --------------------------------------------------------------------------- #
def build_payments(wb: Workbook):
    ws = wb.create_sheet("Payment Summary")
    headers = ["Milestone", "Description", "Amount", "Currency", "Status", "Due Date", "Paid Date"]
    _autosize(ws, [18, 46, 16, 10, 12, 14, 14])
    _banner(ws, len(headers), D.PROJECT_NAME + " — Payment Summary",
            f"Client: {D.CLIENT}  |  Total agreed cost: {D.TOTAL_PROJECT_COST:,.0f} {D.CURRENCY}")

    header_row = 4
    _header_row(ws, header_row, headers)

    r = header_row + 1
    for p in D.PAYMENTS:
        vals = [
            p["milestone"], p["description"], p["amount"], p["currency"],
            p["status"],
            p["due_date"].isoformat() if p["due_date"] else "",
            p["paid_date"].isoformat() if p["paid_date"] else "",
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.alignment = LEFT if col == 2 else CENTER
            cell.border = BORDER
            if col == 3:
                cell.number_format = "#,##0"
            if r % 2 == 0:
                cell.fill = STRIPE_FILL
        r += 1

    # Total row
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=3, value=D.TOTAL_PROJECT_COST).font = TOTAL_FONT
    ws.cell(row=total_row, column=3).number_format = "#,##0"
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = TOTAL_FILL
        c.border = BORDER
        if col not in (1, 3):
            c.value = c.value or ""

    paid = sum(p["amount"] for p in D.PAYMENTS if p["status"] == "Paid")
    pending = D.TOTAL_PROJECT_COST - paid
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value="Paid so far").font = Font(bold=True, color=INK)
    ws.cell(row=summary_row, column=2, value=f"{paid:,.0f} {D.CURRENCY}").font = BODY_FONT
    ws.cell(row=summary_row + 1, column=1, value="Remaining").font = Font(bold=True, color=INK)
    ws.cell(row=summary_row + 1, column=2, value=f"{pending:,.0f} {D.CURRENCY}").font = BODY_FONT

    _status_fill_rules(ws, "E", header_row + 1, total_row - 1)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:G{total_row - 1}"


def main():
    wb = Workbook()
    build_overview(wb)
    build_progress(wb)
    build_payments(wb)
    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")


if __name__ == "__main__":
    main()
